"""
Configuration: allowed dropdown values, model definitions, defaults.
Dropdown values are loaded dynamically from the template at runtime,
but these serve as fallback / validation reference.
"""

MODELS = {
    "opus": "claude-opus-4-20250514",
    "sonnet": "claude-sonnet-4-20250514",
}
DEFAULT_MODEL = "opus"
DEFAULT_BATCH_SIZE = 20
API_URL = "https://api.anthropic.com/v1/messages"
MAX_OUTPUT_TOKENS = 8000


def load_reference_values(template_path):
    """Extract allowed dropdown values from REFERENCE tab of template."""
    import openpyxl
    wb = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    ref = wb["REFERENCE"]

    rows = list(ref.iter_rows(values_only=True))
    wb.close()

    def col_values(col_idx):
        vals = []
        for row in rows[1:]:
            if col_idx < len(row) and row[col_idx] is not None:
                vals.append(str(row[col_idx]).strip())
        return vals

    # REFERENCE layout (0-indexed): col 5=Property, 6=Competitor, 7=Type, 14=Category
    properties = col_values(5)
    competitors = col_values(6)
    types = col_values(7)
    categories = col_values(14)

    return {
        "properties": [v for v in properties if v],
        "competitors": [v for v in competitors if v],
        "types": [v for v in types if v],
        "categories": [v for v in categories if v],
    }


def load_data_headers(template_path):
    """Return ordered list of Data tab column headers."""
    import openpyxl
    wb = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    ws = wb["Data"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    wb.close()
    return headers
