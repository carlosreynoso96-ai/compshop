"""
Stage 5: Write validated offers to a new Excel workbook based on template.
Handles date formatting, formulas, and Step 8/9 compliance.
"""

import os
from datetime import date, datetime
from pathlib import Path

import openpyxl


HEADER_ORDER = [
    "Property", "Competitor", "StartDate", "EndDate", "Type", "Segment",
    "OfferAmt", "PrizeAmtMin", "PrizeAmtMax", "PromotionMonth",
    "PromotionYear", "Category", "Name", "Description", "Source",
]


def parse_date(val):
    """Parse a date string (YYYY-MM-DD) into a date object."""
    if isinstance(val, (date, datetime)):
        return val if isinstance(val, date) else val.date()
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%dT%H:%M:%S"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    return None


def find_insert_row(ws):
    """
    Per Step 8.3: find first row where ALL required cols are blank.
    Required cols (0-indexed): 0,1,2,3,4,5,6,9,10,11,12,13,14
    """
    required_cols = [0, 1, 2, 3, 4, 5, 6, 9, 10, 11, 12, 13, 14]
    row_num = 2  # start checking from row 2 (after header)
    while True:
        all_blank = True
        for col_idx in required_cols:
            cell_val = ws.cell(row=row_num, column=col_idx + 1).value
            if cell_val is not None and str(cell_val).strip() != "":
                all_blank = False
                break
        if all_blank:
            return row_num
        row_num += 1
        if row_num > 100000:
            return row_num


def write_offers(template_path, offers, output_dir=None):
    """
    Write offers to new Excel workbook based on template.
    Returns (output_path, row_count, first_row, last_row).
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Data"]

    insert_row = find_insert_row(ws)

    for i, offer in enumerate(offers):
        r = insert_row + i

        ws.cell(row=r, column=1, value=offer.get("Property", ""))
        ws.cell(row=r, column=2, value=offer.get("Competitor", ""))

        # StartDate
        sd = parse_date(offer.get("StartDate"))
        c_start = ws.cell(row=r, column=3, value=sd)
        c_start.number_format = "yyyy-mm-dd"

        # EndDate
        ed = parse_date(offer.get("EndDate"))
        c_end = ws.cell(row=r, column=4, value=ed)
        c_end.number_format = "yyyy-mm-dd"

        ws.cell(row=r, column=5, value=offer.get("Type", ""))
        ws.cell(row=r, column=6, value=offer.get("Segment", ""))
        ws.cell(row=r, column=7, value=offer.get("OfferAmt", 0))
        ws.cell(row=r, column=8, value=offer.get("PrizeAmtMin", 0))
        ws.cell(row=r, column=9, value=offer.get("PrizeAmtMax", 0))

        # PromotionMonth = TEXT formula
        ws.cell(row=r, column=10, value=f'=TEXT(C{r}, "mmm")')

        # PromotionYear
        if sd:
            ws.cell(row=r, column=11, value=sd.year)
        else:
            ws.cell(row=r, column=11, value="")

        ws.cell(row=r, column=12, value=offer.get("Category", ""))
        ws.cell(row=r, column=13, value=offer.get("Name", ""))
        ws.cell(row=r, column=14, value=offer.get("Description", ""))
        ws.cell(row=r, column=15, value=offer.get("Source", ""))

    # Clear filters, unhide rows
    ws.auto_filter.ref = None
    for rd in ws.row_dimensions.values():
        rd.hidden = False

    # Determine output filename
    # Try to get month from first offer's StartDate
    month_str = ""
    if offers:
        sd = parse_date(offers[0].get("StartDate"))
        if sd:
            month_str = sd.strftime("%Y-%m")
    if not month_str:
        month_str = datetime.now().strftime("%Y-%m")

    output_name = f"MasterCompShopData_{month_str}.xlsx"
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, output_name)
    else:
        output_path = output_name

    wb.save(output_path)

    last_row = insert_row + len(offers) - 1 if offers else insert_row

    return output_path, len(offers), insert_row, last_row


def verify_output(output_path, expected_count, insert_start):
    """
    Step 9 validation: re-read the file and verify.
    Returns dict with verification results.
    """
    wb = openpyxl.load_workbook(output_path, data_only=True)
    ws = wb["Data"]

    required_cols = [0, 1, 2, 3, 4, 5, 6, 9, 10, 11, 12, 13, 14]
    occupied = 0
    first_occupied = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        vals = [row[c].value for c in required_cols]
        if any(v is not None and str(v).strip() != "" for v in vals):
            occupied += 1
            if first_occupied is None:
                first_occupied = row[0].row

    # Preview last 2 rows
    last_row = insert_start + expected_count - 1 if expected_count > 0 else insert_start
    preview = []
    for check_row in range(max(insert_start, last_row - 1), last_row + 1):
        if check_row > ws.max_row:
            break
        vals = {
            "row": check_row,
            "Property": ws.cell(row=check_row, column=1).value,
            "Competitor": ws.cell(row=check_row, column=2).value,
            "StartDate": str(ws.cell(row=check_row, column=3).value),
            "EndDate": str(ws.cell(row=check_row, column=4).value),
            "Type": ws.cell(row=check_row, column=5).value,
            "Segment": ws.cell(row=check_row, column=6).value,
            "OfferAmt": ws.cell(row=check_row, column=7).value,
            "Category": ws.cell(row=check_row, column=12).value,
            "Name": ws.cell(row=check_row, column=13).value,
        }
        preview.append(vals)

    wb.close()

    return {
        "occupied_rows": occupied,
        "expected_rows": expected_count,
        "first_occupied_row": first_occupied,
        "last_row": last_row,
        "count_match": occupied == expected_count,
        "placement_ok": first_occupied == insert_start if first_occupied else True,
        "preview": preview,
    }
